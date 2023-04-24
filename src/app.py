import os
import requests
import socket
import openpyxl
from flask import Flask,jsonify,request,render_template


app = Flask(__name__)

def hostinfo():
    hostname = socket.gethostname()
    print("HOST NAME IS: ",hostname)
    ip = socket.gethostbyname(hostname)
    return(hostname,ip)

@app.route("/health")
def health():
    return jsonify(status="UP")

@app.route("/",methods=["GET","POST"])
def hello():
    h_name,i_addr = hostinfo()
    if request.method == "GET":
        return render_template("index.html",hostname=h_name,ip=i_addr,message="")
    elif request.method == "POST":
        try:
            wb = openpyxl.load_workbook('watched.xlsx')
            print("We in here !!!!")
            ws = wb.active
            m_row, m_col = ws.max_row,ws.max_column
            print('Total number of rows: '+str(m_row)+'. And total number of columns: '+str(m_col))
            ws.cell(row=m_row+1,column=1,value=request.form.get("videoTitle"))
            ws.cell(row=m_row+1,column=2,value=request.form.get("videoLink"))
            ws.cell(row=m_row+1,column=3,value=request.form.get("videoLength"))
            ws.cell(row=m_row+1,column=4,value=request.form.get("playbackSpeed"))
            wb.save('watched.xlsx')
            wb.close()
            return render_template("index.html",hostname=h_name,ip=i_addr,message="Data written Successfully !")
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            # This is only for Mac and Linux OS paths - for windows, need to change
            wb.save(os.getcwd()+"/watched.xlsx")
            ws = wb.active
            ws.cell(row=1,column=1,value="Video Title")
            ws.cell(row=1,column=2,value="URL")
            ws.cell(row=1,column=3,value="Length (min)")
            ws.cell(row=1,column=4,value="Playback speed")
            ws.cell(row=2,column=1,value=request.form.get("videoTitle"))
            ws.cell(row=2,column=2,value=request.form.get("videoLink"))
            ws.cell(row=2,column=3,value=request.form.get("videoLength"))
            ws.cell(row=2,column=4,value=request.form.get("playbackSpeed"))
            wb.save('watched.xlsx')
            wb.close()
            return render_template("index.html",hostname=h_name,ip=i_addr,message="Data written Successfully !")
    else:
        return jsonify(error="Method Unrecognized; not GET or POST")

@app.route("/details")
def details():
    if request.method == "GET":
        h_name,i_addr = hostinfo()
        return render_template("index.html",hostname=h_name,ip=i_addr)

app.run(debug=True)